from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum, Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
import calendar
import uuid
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from .models import Expense, Category, PaymentMethod, PaymentType, MonthlySummary
from .forms import ExpenseForm, ExpenseFilterForm
from .serializers import ExpenseSerializer, CategorySerializer, PaymentMethodSerializer, PaymentTypeSerializer
from accounts.models import CustomUser

def get_first_monday(year, month):
    """Calculate the first Monday of a given month, handling special case when 1st is Monday"""
    first_day = datetime(year, month, 1).date()
    if first_day.weekday() == 0:  # Monday
        return first_day + timedelta(days=7)  # Next Monday (8th)
    else:
        days_to_monday = (7 - first_day.weekday()) % 7
        return first_day + timedelta(days=days_to_monday)

@login_required
def expense_list(request):
    """Vista de lista de gastos con filtros y paginación"""
    # Obtener parámetros de filtro
    form = ExpenseFilterForm(request.GET)
    
    # Filtrar gastos
    expenses = Expense.objects.all()
    
    # Aplicar filtros
    if form.is_valid():
        # Filtro por fecha
        if form.cleaned_data.get('date_from'):
            expenses = expenses.filter(date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            expenses = expenses.filter(date__lte=form.cleaned_data['date_to'])
        
        # Filtro por categoría
        if form.cleaned_data.get('category'):
            expenses = expenses.filter(category=form.cleaned_data['category'])
        
        # Filtro por método de pago
        if form.cleaned_data.get('payment_method'):
            expenses = expenses.filter(payment_method=form.cleaned_data['payment_method'])
        
        # Filtro por tipo de pago
        if form.cleaned_data.get('payment_type'):
            expenses = expenses.filter(payment_type=form.cleaned_data['payment_type'])
        
        # Filtro por usuario
        if form.cleaned_data.get('user'):
            expenses = expenses.filter(user=form.cleaned_data['user'])
        
        # Filtro por crédito
        if form.cleaned_data.get('is_credit') is not None:
            is_credit_value = form.cleaned_data['is_credit']
            if is_credit_value == 'True':
                expenses = expenses.filter(is_credit=True)
            elif is_credit_value == 'False':
                expenses = expenses.filter(is_credit=False)
            # Si es string vacío o None, no filtrar
        
        # Filtro por monto mínimo
        if form.cleaned_data.get('min_amount'):
            expenses = expenses.filter(amount__gte=form.cleaned_data['min_amount'])
        
        # Filtro por monto máximo
        if form.cleaned_data.get('max_amount'):
            expenses = expenses.filter(amount__lte=form.cleaned_data['max_amount'])
        
        # Filtro de búsqueda
        if form.cleaned_data.get('search'):
            search_term = form.cleaned_data['search']
            expenses = expenses.filter(
                Q(name__icontains=search_term) | 
                Q(description__icontains=search_term)
            )
    
    # Ordenamiento
    sort_order = form.cleaned_data.get('sort_order', 'newest')
    if sort_order == 'oldest':
        expenses = expenses.order_by('date', 'id')
    else:  # newest por defecto
        expenses = expenses.order_by('-date', '-id')
    
    # Si no hay filtros aplicados, mostrar solo el mes actual
    if not any([form.cleaned_data.get('date_from'), form.cleaned_data.get('date_to'), 
                form.cleaned_data.get('category'), form.cleaned_data.get('payment_method'),
                form.cleaned_data.get('payment_type'), form.cleaned_data.get('user'),
                form.cleaned_data.get('is_credit'), form.cleaned_data.get('min_amount'),
                form.cleaned_data.get('max_amount'), form.cleaned_data.get('search')]):
        
        # Filtrar por mes actual
        today = timezone.now().date()
        first_day = today.replace(day=1)
        if today.month == 12:
            last_day = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        
        expenses = expenses.filter(date__range=[first_day, last_day])
        period_display = f"{first_day.strftime('%B %Y')}"
    else:
        # Mostrar período de filtros aplicados
        if form.cleaned_data.get('date_from') and form.cleaned_data.get('date_to'):
            period_display = f"Desde {form.cleaned_data['date_from'].strftime('%d/%m/%Y')} hasta {form.cleaned_data['date_to'].strftime('%d/%m/%Y')}"
        elif form.cleaned_data.get('date_from'):
            period_display = f"Desde {form.cleaned_data['date_from'].strftime('%d/%m/%Y')}"
        elif form.cleaned_data.get('date_to'):
            period_display = f"Hasta {form.cleaned_data['date_to'].strftime('%d/%m/%Y')}"
        else:
            period_display = "Todos los períodos"
    
    # Calcular total
    total_amount = expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    # Mostrar usuario filtrado
    user_filter_display = None
    if form.cleaned_data.get('user'):
        user_filter_display = form.cleaned_data['user'].username
    
    # Paginación
    paginator = Paginator(expenses, 25)  # 25 gastos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'expenses': page_obj,
        'form': form,
        'total_amount': total_amount,
        'period_display': period_display,
        'user_filter_display': user_filter_display,
    }
    
    return render(request, 'finances/expense_list.html', context)

@login_required
def expense_create(request):
    """Crear un nuevo gasto"""
    if request.method == 'POST':
        print(f"DEBUG: POST data received: {request.POST}")
        form = ExpenseForm(request.POST)
        print(f"DEBUG: Form is valid: {form.is_valid()}")
        if not form.is_valid():
            print(f"DEBUG: Form errors: {form.errors}")
            print(f"DEBUG: Non-field errors: {form.non_field_errors()}")

        if form.is_valid():
            print(f"DEBUG: Form cleaned data: {form.cleaned_data}")
            expense = form.save(commit=False)
            expense.user = request.user
            
            # Manejar lógica de crédito
            if expense.payment_method.name == 'credito':
                # Get installments from POST data since we're using a custom input field
                installments_str = request.POST.get('installments', '1')
                try:
                    installments = int(installments_str)
                except (ValueError, TypeError):
                    installments = 1

                print(f"DEBUG: Processing credit expense - total_amount: {expense.amount}, installments: {installments}")
                expense.is_credit = True
                expense.total_credit_amount = expense.amount  # Use the entered amount as total credit amount
                expense.installments = installments
                expense.current_installment = 0
                expense.remaining_amount = expense.total_credit_amount
                # Cuota 0/N has amount = 0
                expense.amount = 0
                # Update name and description for Cuota 0/N
                original_name = expense.name
                expense.name = f"{original_name} - Cuota 0/{installments}"
                expense.description = f"Monto total={expense.total_credit_amount} Cantidad de cuotas={installments}"

                # Crear grupo de crédito
                import uuid
                credit_group_id = str(uuid.uuid4())
                expense.credit_group_id = credit_group_id

                # Guardar Cuota 0/N
                expense.save()
                print(f"DEBUG: Saved Cuota 0/{installments} with amount 0 on {expense.date}")

                # Crear cuotas futuras (1/N to N/N)
                if expense.installments >= 1:
                    amount_per_installment = expense.total_credit_amount / expense.installments
                    print(f"DEBUG: Creating {expense.installments} installments of {amount_per_installment} each")

                    # Start from the next month
                    start_year = expense.date.year
                    start_month = expense.date.month + 1
                    if start_month > 12:
                        start_month = 1
                        start_year += 1

                    for i in range(1, expense.installments + 1):
                        # Calculate the month for this installment
                        installment_month = start_month + (i - 1)
                        installment_year = start_year
                        while installment_month > 12:
                            installment_month -= 12
                            installment_year += 1

                        # Get first Monday of that month
                        installment_date = get_first_monday(installment_year, installment_month)
                        print(f"DEBUG: Calculated first Monday for installment {i}: {installment_date} (month: {installment_month}/{installment_year})")

                        # Crear cuota
                        installment_expense = Expense.objects.create(
                            user=request.user,
                            date=installment_date,
                            name=f"{original_name} - Cuota {i}/{installments}",
                            amount=amount_per_installment,
                            category=expense.category,
                            payment_method=expense.payment_method,
                            payment_type=expense.payment_type,
                            description=f"Cuota {i} de {expense.installments} - {expense.description or ''}",
                            is_credit=True,
                            total_credit_amount=expense.total_credit_amount,
                            installments=expense.installments,
                            current_installment=i,
                            remaining_amount=expense.total_credit_amount - (amount_per_installment * i),
                            credit_group_id=credit_group_id
                        )
                        print(f"DEBUG: Created installment {i}/{installments} for {installment_date} with amount {amount_per_installment}")

                messages.success(request, f'Gasto a crédito creado exitosamente. Se crearon {expense.installments + 1} cuotas (0/{expense.installments} + {expense.installments} cuotas).')
            else:
                expense.is_credit = False
                expense.save()
                messages.success(request, 'Gasto creado exitosamente.')
            
            return redirect('finances:expense_list')
    else:
        form = ExpenseForm()
    
    context = {
        'form': form,
        'title': 'Nuevo Gasto'
    }
    
    return render(request, 'finances/expense_form.html', context)

@login_required
def expense_edit(request, pk):
    """Editar un gasto existente"""
    expense = get_object_or_404(Expense, pk=pk)
    
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            old_is_credit = expense.is_credit
            old_credit_group_id = expense.credit_group_id
            old_installments = expense.installments
            old_total_amount = expense.total_credit_amount
            
            # Actualizar el gasto
            expense = form.save(commit=False)
            
            # Manejar cambios en crédito
            if expense.payment_method.name == 'credito':
                expense.is_credit = True
                expense.total_credit_amount = form.cleaned_data.get('total_credit_amount', 0)
                expense.installments = form.cleaned_data.get('installments', 1)
                # Calculate the amount for the first installment if this is the first installment
                if expense.current_installment == 0:
                    expense.amount = expense.total_credit_amount / expense.installments
                
                # Si es un gasto a crédito existente y cambió algo importante
                if old_is_credit and old_credit_group_id:
                    # Si cambió el número de cuotas o el monto total, recrear todas las cuotas
                    if (old_installments != expense.installments or 
                        old_total_amount != expense.total_credit_amount):
                        
                        # Eliminar cuotas existentes
                        Expense.objects.filter(credit_group_id=old_credit_group_id).exclude(pk=expense.pk).delete()
                        
                        # Crear nuevas cuotas
                        if expense.installments >= 1:
                            amount_per_installment = expense.total_credit_amount / expense.installments
                            # Update Cuota 0/N
                            expense.amount = 0
                            original_name = expense.name.split(' - Cuota 0/')[0]  # Extract original name
                            expense.name = f"{original_name} - Cuota 0/{expense.installments}"
                            expense.description = f"Monto total={expense.total_credit_amount} Cantidad de cuotas={expense.installments}"
                            expense.save()

                            # Start from the next month
                            start_year = expense.date.year
                            start_month = expense.date.month + 1
                            if start_month > 12:
                                start_month = 1
                                start_year += 1

                            for i in range(1, expense.installments + 1):
                                # Calculate the month for this installment
                                installment_month = start_month + (i - 1)
                                installment_year = start_year
                                while installment_month > 12:
                                    installment_month -= 12
                                    installment_year += 1

                                # Get first Monday of that month
                                installment_date = get_first_monday(installment_year, installment_month)
                                print(f"DEBUG: Recreated first Monday for installment {i}: {installment_date}")

                                # Crear cuota
                                installment_expense = Expense.objects.create(
                                    user=request.user,
                                    date=installment_date,
                                    name=f"{original_name} - Cuota {i}/{expense.installments}",
                                    amount=amount_per_installment,
                                    category=expense.category,
                                    payment_method=expense.payment_method,
                                    payment_type=expense.payment_type,
                                    description=f"Cuota {i} de {expense.installments} - {expense.description or ''}",
                                    is_credit=True,
                                    total_credit_amount=expense.total_credit_amount,
                                    installments=expense.installments,
                                    current_installment=i,
                                    remaining_amount=expense.total_credit_amount - (amount_per_installment * i),
                                    credit_group_id=old_credit_group_id
                                )
                                print(f"DEBUG: Recreated installment {i}/{expense.installments} for {installment_date}")
                        
                        messages.success(request, f'Gasto a crédito actualizado. Se recrearon {expense.installments + 1} cuotas (0/{expense.installments} + {expense.installments} cuotas).')
                    else:
                        # Solo actualizar campos básicos
                        expense.save()
                        messages.success(request, 'Gasto a crédito actualizado exitosamente.')
                else:
                    # Nuevo gasto a crédito
                    import uuid
                    credit_group_id = str(uuid.uuid4())
                    expense.credit_group_id = credit_group_id
                    expense.current_installment = 0
                    expense.remaining_amount = expense.total_credit_amount
                    # Cuota 0/N has amount = 0
                    expense.amount = 0
                    # Update name and description for Cuota 0/N
                    original_name = expense.name
                    expense.name = f"{original_name} - Cuota 0/{expense.installments}"
                    expense.description = f"Monto total={expense.total_credit_amount} Cantidad de cuotas={expense.installments}"
                    expense.save()

                    # Crear cuotas futuras (1/N to N/N)
                    if expense.installments >= 1:
                        amount_per_installment = expense.total_credit_amount / expense.installments
                        print(f"DEBUG: Creating {expense.installments} installments of {amount_per_installment} each")

                        # Start from the next month
                        start_year = expense.date.year
                        start_month = expense.date.month + 1
                        if start_month > 12:
                            start_month = 1
                            start_year += 1

                        for i in range(1, expense.installments + 1):
                            # Calculate the month for this installment
                            installment_month = start_month + (i - 1)
                            installment_year = start_year
                            while installment_month > 12:
                                installment_month -= 12
                                installment_year += 1

                            # Get first Monday of that month
                            installment_date = get_first_monday(installment_year, installment_month)
                            print(f"DEBUG: Calculated first Monday for installment {i}: {installment_date}")

                            # Crear cuota
                            installment_expense = Expense.objects.create(
                                user=request.user,
                                date=installment_date,
                                name=f"{original_name} - Cuota {i}/{expense.installments}",
                                amount=amount_per_installment,
                                category=expense.category,
                                payment_method=expense.payment_method,
                                payment_type=expense.payment_type,
                                description=f"Cuota {i} de {expense.installments} - {expense.description or ''}",
                                is_credit=True,
                                total_credit_amount=expense.total_credit_amount,
                                installments=expense.installments,
                                current_installment=i,
                                remaining_amount=expense.total_credit_amount - (amount_per_installment * i),
                                credit_group_id=credit_group_id
                            )
                            print(f"DEBUG: Created installment {i}/{expense.installments} for {installment_date}")

                        messages.success(request, f'Gasto a crédito creado exitosamente. Se crearon {expense.installments + 1} cuotas (0/{expense.installments} + {expense.installments} cuotas).')
            else:
                # No es crédito
                expense.is_credit = False
                expense.total_credit_amount = None
                expense.installments = None
                expense.current_installment = None
                expense.remaining_amount = None
                expense.credit_group_id = None
                expense.save()
                
                # Si antes era crédito, eliminar cuotas relacionadas
                if old_is_credit and old_credit_group_id:
                    Expense.objects.filter(credit_group_id=old_credit_group_id).exclude(pk=expense.pk).delete()
                
                messages.success(request, 'Gasto actualizado exitosamente.')
            
            return redirect('finances:expense_list')
    else:
        form = ExpenseForm(instance=expense)
    
    context = {
        'form': form,
        'expense': expense
    }
    
    return render(request, 'finances/expense_form.html', context)

@login_required
def expense_delete(request, pk):
    """Eliminar un gasto"""
    expense = get_object_or_404(Expense, pk=pk)
    
    if request.method == 'POST':
        # Si es un gasto a crédito, eliminar todas las cuotas relacionadas
        if expense.is_credit and expense.credit_group_id:
            related_expenses = Expense.objects.filter(credit_group_id=expense.credit_group_id)
            count = related_expenses.count()
            related_expenses.delete()
            messages.success(request, f'Gasto a crédito y {count-1} cuotas relacionadas eliminadas exitosamente.')
        else:
            expense.delete()
            messages.success(request, 'Gasto eliminado exitosamente.')
        
        return redirect('finances:expense_list')
    
    context = {
        'expense': expense
    }
    
    return render(request, 'finances/expense_confirm_delete.html', context)

@login_required
def expense_detail(request, pk):
    """Ver detalle de un gasto"""
    expense = get_object_or_404(Expense, pk=pk)
    
    # Obtener gastos relacionados si es crédito
    related_expenses = []
    if expense.is_credit and expense.credit_group_id:
        related_expenses = Expense.objects.filter(
            credit_group_id=expense.credit_group_id
        ).exclude(pk=expense.pk).order_by('current_installment')
    
    context = {
        'expense': expense,
        'related_expenses': related_expenses
    }
    
    return render(request, 'finances/expense_detail.html', context)

@login_required
def dashboard_finances(request):
    """Dashboard financiero con estadísticas"""
    current_month = timezone.now().month
    current_year = timezone.now().year
    
    # Gastos del mes actual
    current_month_expenses = Expense.objects.filter(
        date__year=current_year,
        date__month=current_month
    )
    
    # Total del mes
    month_total = current_month_expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    # Gastos por categoría del mes
    expenses_by_category = current_month_expenses.values('category__name').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Créditos pendientes
    credit_pending = Expense.objects.filter(
        is_credit=True
    ).aggregate(total=Sum('remaining_amount'))['total'] or 0
    
    # Próximas cuotas (próximo mes)
    next_month = current_month + 1 if current_month < 12 else 1
    next_year = current_year if current_month < 12 else current_year + 1
    
    upcoming_installments = Expense.objects.filter(
        is_credit=True,
        date__year=next_year,
        date__month=next_month
    ).order_by('date')
    
    # Historial de últimos 6 meses
    months_data = []
    for i in range(6):
        month = current_month - i
        year = current_year
        if month < 1:
            month += 12
            year -= 1
        
        month_expenses = Expense.objects.filter(
            date__year=year,
            date__month=month
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        months_data.append({
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month][:3],
            'total': month_expenses
        })
    
    months_data.reverse()
    
    context = {
        'month_total': month_total,
        'expenses_by_category': expenses_by_category,
        'credit_pending': credit_pending,
        'upcoming_installments': upcoming_installments,
        'months_data': months_data,
        'current_month': current_month,
        'current_year': current_year
    }
    return render(request, 'finances/dashboard.html', context)

@login_required
def export_expenses(request):
    """Exportar gastos a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    # Obtener gastos filtrados
    expenses = Expense.objects.all()
    
    # Aplicar filtros si existen
    form = ExpenseFilterForm(request.GET)
    if form.is_valid():
        # Aplicar los mismos filtros que en la vista de lista
        pass
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"
    
    # Estilos
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Headers
    headers = ['Fecha', 'Usuario', 'Nombre', 'Monto', 'Categoría', 'Método de Pago', 'Tipo', 'Es Crédito', 'Cuotas', 'Descripción']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Datos
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=expense.date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=expense.user.username)
        ws.cell(row=row, column=3, value=expense.name)
        ws.cell(row=row, column=4, value=float(expense.amount))
        ws.cell(row=row, column=5, value=expense.category.name)
        ws.cell(row=row, column=6, value=expense.payment_method.get_name_display())
        ws.cell(row=row, column=7, value=expense.payment_type.get_name_display())
        ws.cell(row=row, column=8, value='Sí' if expense.is_credit else 'No')
        ws.cell(row=row, column=9, value=f"{expense.current_installment}/{expense.installments}" if expense.is_credit else '')
        ws.cell(row=row, column=10, value=expense.description or '')
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=gastos_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response

def get_payment_types(request):
    """Vista AJAX para obtener tipos de pago según el método seleccionado"""
    payment_method_param = request.GET.get('payment_method')
    if payment_method_param:
        # Filter by payment_method id (since form now uses pk)
        payment_types = PaymentType.objects.filter(
            payment_method_id=payment_method_param
        ).order_by('-is_default', 'name')

        data = [{'id': pt.id, 'name': pt.get_name_display(), 'is_default': pt.is_default} for pt in payment_types]
        return JsonResponse({'payment_types': data}, safe=False)

    return JsonResponse({'payment_types': []}, safe=False)


class IsOwnerOrReadOnly(permissions.BasePermission):
    """
    Custom permission to only allow owners of an object to edit it.
    """
    def has_object_permission(self, request, view, obj):
        # Read permissions are allowed to any request,
        # so we'll always allow GET, HEAD or OPTIONS requests.
        if request.method in permissions.SAFE_METHODS:
            return True

        # Write permissions are only allowed to the owner of the expense.
        return obj.user == request.user


class ExpenseViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Expense model with filtering and user ownership permissions.
    """
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrReadOnly]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['user', 'date', 'category', 'payment_method', 'payment_type', 'is_credit']
    ordering_fields = ['date', 'amount', 'created_at']
    ordering = ['-date']

    def get_queryset(self):
        """
        Filter queryset to only show expenses for the authenticated user,
        unless user_id is specified in query params (for admin access).
        """
        queryset = Expense.objects.all()

        # Filter by user - authenticated user can only see their own expenses
        # unless they specify user_id in query params
        user_id = self.request.query_params.get('user_id')
        if user_id:
            # Allow filtering by specific user_id if provided
            queryset = queryset.filter(user_id=user_id)
        else:
            # Default: only show authenticated user's expenses
            queryset = queryset.filter(user=self.request.user)

        # Date range filtering
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        return queryset

    def perform_create(self, serializer):
        """
        Set the user to the authenticated user when creating expenses.
        """
        serializer.save(user=self.request.user)


class CategoryViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only ViewSet for Category model.
    """
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]


class PaymentMethodViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only ViewSet for PaymentMethod model.
    """
    queryset = PaymentMethod.objects.all()
    serializer_class = PaymentMethodSerializer
    permission_classes = [permissions.IsAuthenticated]


class PaymentTypeViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only ViewSet for PaymentType model.
    """
    queryset = PaymentType.objects.all()
    serializer_class = PaymentTypeSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['payment_method']
