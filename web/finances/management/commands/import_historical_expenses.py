from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from django.utils import timezone
from datetime import datetime, timedelta
import calendar
import openpyxl
import os
import uuid
from finances.models import Expense, Category, PaymentMethod, PaymentType

class Command(BaseCommand):
    help = 'Importa gastos históricos desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='/gastos/Gastos-Telegram.XLSX',
            help='Ruta al archivo Excel con gastos históricos'
        )
        parser.add_argument(
            '--user',
            type=str,
            default='admin',
            help='Usuario que será asignado a los gastos importados'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        username = options['user']

        # Verificar que el archivo existe
        if not os.path.exists(file_path):
            self.stdout.write(
                self.style.ERROR(f'❌ Archivo no encontrado: {file_path}')
            )
            return

        # Obtener usuario
        User = get_user_model()
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            self.stdout.write(
                self.style.ERROR(f'❌ Usuario no encontrado: {username}')
            )
            return

        self.stdout.write(f'🔄 Importando gastos para usuario: {user.username}')

        # Cargar el archivo Excel
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'❌ Error al abrir archivo Excel: {e}')
            )
            return

        # Mapeo de tipos de pago
        payment_type_mapping = {
            'crédito': 'credito',
            'débito': 'debito',
            'efectivo': 'efectivo'
        }

        # Mapeo inteligente de categorías
        category_mapping = {
            # Alimentos
            'carne': 'Carnicería',
            'pollo': 'Carnicería',
            'fiambre': 'Carnicería',
            'matambre': 'Carnicería',
            'súper': 'Supermercado',
            'super': 'Supermercado',
            'la anonima': 'Supermercado',
            'la anónima': 'Supermercado',
            'la ecologica': 'Supermercado',
            'eden': 'Supermercado',
            'miga': 'Panadería',
            'pan': 'Panadería',
            'papitas': 'Supermercado',
            'chocolate': 'Supermercado',
            'dulce': 'Supermercado',
            'helado': 'Supermercado',
            'empanadas': 'Supermercado',
            'gula': 'Supermercado',
            'jarabe': 'Medicamentos',
            'medicamento': 'Medicamentos',
            'medicina': 'Medicamentos',

            # Transporte
            'combustible': 'Combustible',
            'nafta': 'Combustible',
            'gasolina': 'Combustible',
            'seguro': 'Mantenimiento Auto',
            'kangoo': 'Mantenimiento Auto',
            'auto': 'Mantenimiento Auto',
            'coche': 'Mantenimiento Auto',

            # Entretenimiento
            'salida': 'Salidas',
            'birreria': 'Salidas',
            'café': 'Salidas',
            'cafe': 'Salidas',
            'cumbal': 'Salidas',
            'baile': 'Cursos',
            'curso': 'Cursos',
            'libro': 'Libros',
            'ingles': 'Cursos',
            'idioma': 'Cursos',

            # Hogar
            'aire': 'Electrodomésticos',
            'microondas': 'Electrodomésticos',
            'cortadora': 'Jardín',
            'arbolito': 'Jardín',
            'jardin': 'Jardín',
            'quinta': 'Jardín',
            'red power': 'Electrónica',
            'turbo': 'Electrodomésticos',
            'ventilator': 'Electrodomésticos',

            # Ropa
            'ropa': 'Otros',
            'zapatillas': 'Otros',
            'zapas': 'Otros',
            'zapatos': 'Otros',
            'vuelta al cole': 'Material Escolar',
            'escolar': 'Material Escolar',
            'colegio': 'Material Escolar',
        }

        total_imported = 0
        total_credits = 0

        # Procesar cada hoja (mes)
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() in ['hoja1', 'sheet1', 'datos']:
                continue  # Saltar hojas genéricas

            self.stdout.write(f'📅 Procesando mes: {sheet_name}')
            sheet = workbook[sheet_name]

            # Buscar encabezados
            headers = []
            header_row = None

            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row and any(cell for cell in row if cell and 'fecha' in str(cell).lower()):
                    headers = [str(cell).lower().strip() if cell else '' for cell in row]
                    header_row = row_num
                    break

            if not headers:
                self.stdout.write(f'  ⚠️  No se encontraron encabezados en {sheet_name}')
                continue

            # Mapear columnas con sinónimos (MEJORA: Detección flexible de encabezados)
            col_synonyms = {
                'fecha': ['fecha', 'date', 'fecha_transaccion'],
                'nombre': ['nombre', 'name', 'gasto'],
                'valor': ['valor', 'amount', 'monto'],
                'tipo': ['tipo', 'type', 'payment_type'],
                'cuota_actual': ['cuota actual', 'current installment', 'cuota_actual'],
                'cuotas_restantes': ['cuotas restantes', 'remaining installments', 'cuotas_restantes']
            }

            col_mapping = {}
            for key, synonyms in col_synonyms.items():
                for i, header in enumerate(headers):
                    if any(syn in header for syn in synonyms):
                        col_mapping[key] = i
                        break

            if not all(key in col_mapping for key in ['fecha', 'nombre', 'valor', 'tipo']):
                self.stdout.write(f'  ⚠️  Columnas requeridas no encontradas en {sheet_name}: {[k for k in ["fecha", "nombre", "valor", "tipo"] if k not in col_mapping]}')
                continue

            # Procesar filas de datos
            for row_num, row in enumerate(sheet.iter_rows(values_only=True, min_row=header_row + 1), header_row + 1):
                if not any(cell for cell in row if cell):
                    continue  # Fila vacía

                try:
                    # Extraer datos
                    fecha_str = str(row[col_mapping['fecha']]).strip()
                    nombre = str(row[col_mapping['nombre']]).strip()
                    valor_str = str(row[col_mapping['valor']]).strip()
                    tipo = str(row[col_mapping['tipo']]).strip().lower()

                    # Validar datos básicos
                    if not fecha_str or not nombre or not valor_str or not tipo:
                        continue

                    # Parsear fecha con fallbacks (MEJORA: Soporte para formatos mixtos)
                    fecha = self.parse_date(fecha_str)
                    if not fecha:
                        self.stdout.write(f'    ⚠️  Fecha inválida: {fecha_str}')
                        continue

                    # Parsear valor con validaciones (MEJORA: Validación de negativos/cero y símbolos extra)
                    valor = self.parse_amount(valor_str)
                    if valor is None or valor <= 0:
                        self.stdout.write(f'    ⚠️  Valor inválido: {valor_str}')
                        continue

                    # Determinar categoría
                    categoria_nombre = 'Otros'  # Por defecto
                    nombre_lower = nombre.lower()

                    for keyword, categoria in category_mapping.items():
                        if keyword in nombre_lower:
                            categoria_nombre = categoria
                            break

                    # Obtener o crear categoría
                    categoria, _ = Category.objects.get_or_create(
                        name=categoria_nombre,
                        defaults={'is_active': True}
                    )

                    # Determinar método de pago
                    if tipo in payment_type_mapping:
                        payment_method_name = payment_type_mapping[tipo]
                        try:
                            payment_method = PaymentMethod.objects.get(name=payment_method_name)
                        except PaymentMethod.DoesNotExist:
                            self.stdout.write(f'    ⚠️  Método de pago no encontrado: {payment_method_name}')
                            continue
                    else:
                        self.stdout.write(f'    ⚠️  Tipo de pago no reconocido: {tipo}')
                        continue

                    # Obtener tipo de pago por defecto
                    try:
                        payment_type = PaymentType.objects.filter(
                            payment_method=payment_method,
                            is_default=True
                        ).first()
                        if not payment_type:
                            payment_type = PaymentType.objects.filter(
                                payment_method=payment_method
                            ).first()
                    except PaymentType.DoesNotExist:
                        self.stdout.write(f'    ⚠️  Tipo de pago no encontrado para: {payment_method_name}')
                        continue

                    # Manejar créditos con correcciones (MEJORA: Eliminación de cuota 0, deduplicación, cálculo correcto)
                    if tipo == 'crédito':
                        cuota_actual = int(row[col_mapping.get('cuota_actual', 0)] or 1)  # Default a 1 si falta
                        cuotas_restantes = int(row[col_mapping.get('cuotas_restantes', 0)] or 0)
                        total_cuotas = cuota_actual + cuotas_restantes

                        if total_cuotas <= 0:
                            self.stdout.write(f'    ⚠️  Invalid installment count: {total_cuotas}')
                            continue

                        # Deduplicación: Verificar si ya existe un grupo similar (MEJORA: Evitar duplicados)
                        existing_credit = Expense.objects.filter(
                            user=user,
                            name=nombre,
                            is_credit=True,
                            total_credit_amount=valor * total_cuotas  # Asumir valor es por cuota
                        ).first()

                        if existing_credit:
                            self.stdout.write(f'    ⚠️  Crédito duplicado detectado: {nombre}, saltando')
                            continue

                        # Crear grupo de crédito sin cuota 0
                        credit_group_id = str(uuid.uuid4())

                        # Calcular monto total asumiendo valor es por cuota
                        total_amount = valor * total_cuotas
                        remaining = total_amount - (valor * (cuota_actual - 1))  # Cálculo correcto de remaining

                        # Crear cuota actual
                        Expense.objects.create(
                            user=user,
                            date=fecha,
                            name=nombre,
                            amount=valor,
                            category=categoria,
                            payment_method=payment_method,
                            payment_type=payment_type,
                            description=f'Importado desde Excel - {sheet_name} - Cuota {cuota_actual}',
                            is_credit=True,
                            total_credit_amount=total_amount,
                            installments=total_cuotas,
                            current_installment=cuota_actual,
                            remaining_amount=remaining,
                            credit_group_id=credit_group_id
                        )

                        # Crear cuotas futuras si hay restantes
                        current_date = self.calculate_next_installment_date(fecha)  # Mejora en fechas
                        for i in range(cuota_actual + 1, total_cuotas + 1):
                            remaining -= valor
                            Expense.objects.create(
                                user=user,
                                date=current_date,
                                name=nombre,
                                amount=valor,
                                category=categoria,
                                payment_method=payment_method,
                                payment_type=payment_type,
                                description=f'Importado desde Excel - {sheet_name} - Cuota {i}',
                                is_credit=True,
                                total_credit_amount=total_amount,
                                installments=total_cuotas,
                                current_installment=i,
                                remaining_amount=max(remaining, 0),  # Evitar negativos
                                credit_group_id=credit_group_id
                            )
                            current_date = self.calculate_next_installment_date(current_date)

                        total_credits += 1
                        self.stdout.write(f'    ✅ Crédito creado: {nombre} - {total_cuotas} cuotas')
                    else:
                        # Gasto normal
                        Expense.objects.create(
                            user=user,
                            date=fecha,
                            name=nombre,
                            amount=valor,
                            category=categoria,
                            payment_method=payment_method,
                            payment_type=payment_type,
                            description=f'Importado desde Excel - {sheet_name}',
                            is_credit=False
                        )
                        self.stdout.write(f'    ✅ Gasto: {nombre} - ${valor:.2f}')

                    total_imported += 1

                except Exception as e:
                    self.stdout.write(f'    ❌ Error procesando fila {row_num}: {e}')
                    continue

        # Resumen final
        self.stdout.write('\n' + '='*50)
        self.stdout.write(
            self.style.SUCCESS(
                f'✅ Importación completada!\n'
                f'📊 Total de gastos importados: {total_imported}\n'
                f'💳 Total de créditos procesados: {total_credits}\n'
                f'👤 Usuario asignado: {user.username}'
            )
        )

        workbook.close()

    # Método auxiliar para parsear fechas con fallbacks (MEJORA)
    def parse_date(self, fecha_str):
        formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']
        for fmt in formats:
            try:
                return datetime.strptime(fecha_str, fmt).date()
            except ValueError:
                continue
        return None

    # Método auxiliar para calcular próxima fecha de cuota (MEJORA: Ajuste para vencimientos)
    def calculate_next_installment_date(self, current_date):
        # Asumir mensual, ajustar al primer lunes si es día 1
        next_date = current_date.replace(month=current_date.month + 1)
        if next_date.day == 1:
            # Si cae en 1ro, mover al primer lunes
            days_to_monday = (7 - next_date.weekday()) % 7
            if days_to_monday == 0:
                days_to_monday = 7
            next_date += timedelta(days=days_to_monday)
        return next_date

    # Método auxiliar para parsear montos con validaciones (MEJORA)
    def parse_amount(self, valor_str):
        # Limpiar símbolos extra
        valor_str = valor_str.replace('$', '').replace(' ', '').strip()
        is_negative = valor_str.startswith('-')
        valor_str = valor_str.lstrip('-')

        # Parsing existente
        if ',' in valor_str and '.' not in valor_str:
            valor = float(valor_str.replace(',', '.'))
        elif ',' in valor_str and '.' in valor_str:
            parts = valor_str.split(',')
            if len(parts) == 2 and len(parts[1]) == 2:
                valor = float(parts[0].replace('.', '') + '.' + parts[1])
            else:
                valor = float(valor_str.replace(',', ''))
        else:
            valor = float(valor_str)

        if is_negative:
            valor = -valor

        # Validaciones
        if valor == 0 or valor < -10000 or valor > 100000:
            return None
        return valor
